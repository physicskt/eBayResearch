import win32com.client
import os
import openpyxl


def close_excel_file(file_name):
    # Excelアプリケーションを取得
    excel = win32com.client.Dispatch("Excel.Application")
    
    # Excelのすべてのワークブックを確認
    for wb in excel.Workbooks:
        # ファイル名が一致するワークブックを強制的に閉じる
        if os.path.basename(wb.FullName) == os.path.basename(file_name):
            print(f"Closing {file_name}...")
            wb.Close(SaveChanges=True)  # SaveChanges=True にすると保存されます
            break
    else:
        # print(f"{file_name} is not open in Excel.")
        pass


def open_visible_excel(file_name):
    import win32com.client

    # Excelを起動
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = True  # Excelを表示

    # Excelファイルを開く
    wb = excel.Workbooks.Open(os.path.abspath(file_name))


def write_to_data(file_name, sheet_name, data):
    # 追加するデータの配列（例）
    # data = [["Name", "Age", "City"], 
    #         ["Alice", 30, "New York"],
    #         ["Bob", 25, "Los Angeles"],
    #         ["Charlie", 35, "Chicago"]]

    close_excel_file(file_name)

    # エクセルファイルを開く（新規作成する場合は None を渡す）
    # file_name = "example.xlsx"
    try:
        wb = openpyxl.load_workbook(file_name)
    except FileNotFoundError:
        wb = openpyxl.Workbook()  # ファイルが存在しない場合、新規作成

    # 新しいシートを追加（"シート名"という名前で）
    # sheet_name = "シート名"
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # シートの全行を削除
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(sheet_name)

    # 配列のデータをシートに書き込む
    for row in data:
        ws.append(row)

    # エクセルファイルを保存
    wb.save(file_name)
    # エクセルファイルを開く
    # open_visible_excel(file_name)


def append_data_to_excel(file_path="scraiping_list.xlsx", sheet_name="Sheet1", data=["AAA","DDD","CCC"], start_row=1, column="B"):
    # # 使用例
    # file_path = "example.xlsx"
    # sheet_name = "Sheet1"
    # data = [10, 20, 30]  # 追加するデータ
    # start_row = 5  # 追加開始行（1からの行番号）
    # column = "B"  # 追加する列（1列目はA列）

    column=column_letter_to_number(column)

    close_excel_file(file_path)
    # Excelファイルを開く
    try:
        wb = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        print(f"{file_path} は存在しません")
        return

    # シートを取得
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        print(f"シート '{sheet_name}' は存在しません")
        return

    # `start_row` 以降の行と指定列以降をリセット（空にする）
    for row in range(start_row, ws.max_row + 1):  # 行は start_row から最終行まで
        for col in range(column, ws.max_column + 1):  # 列は指定された列から最後の列まで
            ws.cell(row=row, column=col, value=None)  # セルを空に設定

    # データを指定した行と列に追加
    for i, value in enumerate(data):
        # `start_row`は1からの行番号なので、適切な行にデータを追加
        ws.cell(row=start_row + i, column=column, value=value)

    # Excelファイルを保存
    wb.save(file_path)


def column_letter_to_number(column_letter:str):
    """列アルファベットを列番号に変換（例: 'A' -> 1, 'B' -> 2）"""
    return openpyxl.utils.column_index_from_string(column_letter)


if __name__ == "__main__":
    # 5行2列のデータを書き込み
    # write_to_data("scraiping_list.xlsx", "1", [["data1","data1","data1","data1","data1"], ["data2","data2","data2","data2","data2"]])
    # open_visible_excel("scraiping_list.xlsx")
    append_data_to_excel(start_row=2, column="C", data=[1,2,3,4])
    open_visible_excel("scraiping_list.xlsx")
