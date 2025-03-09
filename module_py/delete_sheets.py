import openpyxl

def delete_used_sheets(file_name, used_sheets):
    # エクセルファイルを開く
    wb = openpyxl.load_workbook(file_name)

    # 配列に含まれるシート名をループして削除
    for sheet_name in used_sheets:
        if sheet_name in wb.sheetnames:  # シートが存在するかチェック
            del wb[sheet_name]  # シート削除
        else:
            print(f"シート [{sheet_name}] は存在しません。")

    # 変更を保存
    wb.save(file_name)
    print(f"指定されたシートが削除されました。")


def delete_other_sheets(file_name, used_sheets):
    # エクセルファイルを開く
    wb = openpyxl.load_workbook(file_name)

    # すべてのシート名を取得
    all_sheets = wb.sheetnames

    # used_sheetsに含まれていないシートを削除
    for sheet_name in all_sheets:
        if sheet_name not in used_sheets:  # used_sheetsに含まれていないシートを削除
            del wb[sheet_name]
            print(f"シート [{sheet_name}] は削除されました。")

    # 変更を保存
    wb.save(file_name)
    print("指定されたシート以外が削除されました。")


if __name__ == "__main__":
    # 使用例
    used_sheets = ["ワードリスト", "シート1", "シート2", "シート3"]  # 残すシート名
    file_name = "scraiping_list.xlsx"
    delete_other_sheets(file_name, used_sheets)

