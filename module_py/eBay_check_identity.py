import openpyxl
# from difflib import SequenceMatcher
import write_excel

# 関数: 文字列が3文字以上一致するかチェック
def is_similar(str1, str2, min_words=3):
    # str1を空白で分割
    words_in_str1 = str1.split()
    # str2に含まれる単語の数をカウント
    matching_words = sum(1 for word in words_in_str1 if word in str2)
    return matching_words >= min_words

def check_and_calc_sold(file_name="scraiping_list.xlsx", sheet_name="1", similality_check_count=3):
    # エクセルファイルのパス
    # file_name = "example.xlsx"
    pattern1_col = 6
    write_excel.close_excel_file(file_name)
    wb = openpyxl.load_workbook(file_name)
    try:
        ws = wb[sheet_name]  # シート名を指定
        print(f"シート [{sheet_name}] 処理開始")
    except Exception as e:
        print(f"！！！！！シート [{sheet_name}] が存在しません。")
        return

    max_row = ws.max_row  # 最大行数を取得

    pattern1_flags = {}
    # 列リセット
    for i in range(1, max_row + 1):  # 1行目から最終行までループ（ヘッダーを含む）
        ws.cell(row=i, column=pattern1_col).value = None  # E列を空に設定
        ws.cell(row=i, column=pattern1_col+1).value = None  # E列を空に設定
        ws.cell(row=i, column=pattern1_col+2).value = None  # E列を空に設定
        ws.cell(row=i, column=pattern1_col+3).value = None  # E列を空に設定
        pattern1_flags[i] = True

    # ヘッダー入力
    ws.cell(row=1, column=pattern1_col).value = "パターン①該当数"
    ws.cell(row=1, column=pattern1_col+1).value = "パターン①該当行"
    ws.cell(row=1, column=pattern1_col+2).value = "パターン①により処理するか？"
    ws.cell(row=1, column=pattern1_col+3).value = "パターン③該当か？"

    # C列、D列、E列のデータを取得
    for i in range(2, max_row + 1):  # 2行目から最終行までループ（ヘッダーを除外）
        c_value = ws.cell(row=i, column=3).value  # C列の数字
        d_value = ws.cell(row=i, column=4).value  # D列の文字列
        pattern1_rows = ""
        pattern3_judge = False

        try:
            c_value = float(c_value)
        except Exception as e:
            print(e)
            print("！！！！！c_valueは文字列です。数値に変換できません。")
            continue

        # 文字列が空でないかチェック
        if not d_value:
            continue
        
        total = 0
        for j in range(2, max_row + 1):  # 他の行と比較
            other_c_value = ws.cell(row=j, column=3).value  # 他の行のC列の数字
            try:
                other_c_value = float(other_c_value)
            except Exception as e:
                print(e)
            
            other_d_value = ws.cell(row=j, column=4).value  # 他の行のD列の文字列

            if is_similar(d_value, other_d_value, similality_check_count) or (i == j):
                if isinstance(other_c_value, (int, float)):  # 数字であるかチェック
                    total += other_c_value
                    pattern1_rows = pattern1_rows + str(j) + ","

                if pattern1_flags[j] and (j > i):
                    pattern1_flags[j] = False

        if c_value > 2:
            pattern3_judge = True

        pattern1_rows = pattern1_rows[:-1]

        ws.cell(row=i, column=pattern1_col).value = total
        ws.cell(row=i, column=pattern1_col+1).value = pattern1_rows
        ws.cell(row=i, column=pattern1_col+2).value = pattern1_flags[i]
        ws.cell(row=i, column=pattern1_col+3).value = pattern3_judge

    # エクセルファイルを保存
    wb.save(file_name)
    print(f"シート [{sheet_name}] の処理が完了しました。")


if __name__ == "__main__":
    # 各シートに対して処理を実行
    check_and_calc_sold(sheet_name="1")
    check_and_calc_sold(sheet_name="2")
    check_and_calc_sold(sheet_name="3")
    check_and_calc_sold(sheet_name="4")
    check_and_calc_sold(sheet_name="5")
    check_and_calc_sold(sheet_name="6")

    # エクセルファイルを表示（最終的な操作）
    write_excel.open_visible_excel("scraiping_list.xlsx")
